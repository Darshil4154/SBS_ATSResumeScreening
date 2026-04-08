import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from evaluator import CRITERIA_ORDER, CRITERIA_LABELS


BLACK_FILL = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LIGHT_YELLOW_FILL = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
LIGHT_RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

WHITE_FONT = Font(color='FFFFFF', bold=True, size=10)
YELLOW_FONT = Font(color='FFFF00', bold=True, size=10)
BLACK_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def score_fill(score):
    if score >= 8:
        return GREEN_FILL
    elif score >= 5:
        return LIGHT_YELLOW_FILL
    else:
        return LIGHT_RED_FILL


def export_to_excel(candidates, export_dir):
    wb = Workbook()

    # Convert sqlite3.Row objects to dicts
    candidates = [dict(c) for c in candidates]

    # ── Sheet 1: Scoring Matrix ──
    ws1 = wb.active
    ws1.title = "Scoring Matrix"

    # Row 1: merged header
    ws1.merge_cells('C1:H1')
    ws1['C1'] = 'PREFERRED KNOWLEDGE, SKILLS, AND ABILITIES'
    ws1['C1'].font = Font(bold=True, size=12)
    ws1['C1'].alignment = Alignment(horizontal='center')

    # Row 1 A-B: instructions block
    ws1.merge_cells('A1:B1')
    ws1['A1'] = (
        "Candidates in this stage have been determined to meet the minimum qualifications "
        "of the position for which they applied. All candidates must be evaluated against a "
        "common set of criteria to determine their suitability for the next step in the interview process."
    )
    ws1['A1'].font = Font(size=9)
    ws1['A1'].alignment = Alignment(wrap_text=True, vertical='top')

    # Column mapping for row 2 headers
    col_map = [
        ('A', 'APPLICANTS', BLACK_FILL, WHITE_FONT),
        ('B', '', BLACK_FILL, WHITE_FONT),
        ('C', 'Project Management', BLACK_FILL, WHITE_FONT),
        ('D', 'Software Testing/ERP\nImplementation', BLACK_FILL, WHITE_FONT),
        ('E', 'Creating Reports/SQL', BLACK_FILL, WHITE_FONT),
        ('F', 'SIS Business\nProcesses in AR', BLACK_FILL, WHITE_FONT),
        ('G', 'Process Improvement', BLACK_FILL, WHITE_FONT),
        ('H', 'Reconciling Accounts', BLACK_FILL, WHITE_FONT),
        ('I', 'APPLICATION\nSCORE', YELLOW_FILL, BLACK_FONT),
        ('J', 'GAAP & Financial\nMgmt', BLACK_FILL, WHITE_FONT),
        ('K', 'Federal/State\nCompliance', BLACK_FILL, WHITE_FONT),
        ('L', 'Budget & Financial\nAnalysis', BLACK_FILL, WHITE_FONT),
        ('M', 'Communication &\nExec Presence', BLACK_FILL, WHITE_FONT),
        ('N', 'SOP &\nDocumentation', BLACK_FILL, WHITE_FONT),
        ('O', 'Vendor &\nStakeholder Coord', BLACK_FILL, WHITE_FONT),
        ('P', 'Education Match', BLACK_FILL, WHITE_FONT),
        ('Q', 'Experience Match', BLACK_FILL, WHITE_FONT),
        ('R', 'INTERVIEW 1\nSCORE', BLACK_FILL, WHITE_FONT),
        ('S', 'INTERVIEW 2\nSCORE', BLACK_FILL, WHITE_FONT),
        ('T', 'REFERENCE\nSCORE', BLACK_FILL, WHITE_FONT),
        ('U', 'Military Preference\nClaimed', YELLOW_FILL, BLACK_FONT),
        ('V', 'Former Foster Child\nPreference Claimed', YELLOW_FILL, BLACK_FONT),
        ('W', 'TOTAL SCORE', RED_FILL, WHITE_FONT),
    ]

    for col_letter, header, fill, font in col_map:
        cell = ws1[f'{col_letter}2']
        cell.value = header
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            text_rotation=90, horizontal='center', vertical='bottom', wrap_text=True
        )
        cell.border = THIN_BORDER

    ws1.row_dimensions[2].height = 120

    # Score column order mapping (spreadsheet columns)
    # C-H = first 6 criteria, then I=APPLICATION SCORE, J-O = criteria 7-12, P-Q = 13-14
    score_col_order = [
        'project_management',        # C
        'software_testing_erp',      # D
        'reports_sql',               # E
        'sis_business_processes_ar',  # F
        'process_improvement',        # G
        'reconciling_accounts',       # H
        # I = APPLICATION SCORE
        'gaap_financial_mgmt',                # J
        'federal_state_compliance',           # K
        'budget_financial_analysis',          # L
        'communication_executive_presence',   # M
        'sop_documentation',                  # N
        'vendor_stakeholder_coordination',    # O
        'education_match',                    # P
        'experience_match',                   # Q
    ]

    # Data rows
    for idx, c in enumerate(candidates):
        row = idx + 3
        scores = json.loads(c['scores_json']) if c['scores_json'] else {}
        name = c['candidate_name'] or c['filename']

        ws1[f'A{row}'] = name
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'A{row}'].border = THIN_BORDER

        # First 6 criteria → C through H
        for i, key in enumerate(score_col_order[:6]):
            col_letter = get_column_letter(3 + i)  # C=3
            s = scores.get(key, {}).get('score', 0)
            cell = ws1[f'{col_letter}{row}']
            cell.value = s
            cell.fill = score_fill(s)
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # I = APPLICATION SCORE
        app_score = c['application_score'] or 0
        cell = ws1[f'I{row}']
        cell.value = app_score
        cell.fill = YELLOW_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

        # J through Q = criteria 7-14
        for i, key in enumerate(score_col_order[6:]):
            col_letter = get_column_letter(10 + i)  # J=10
            s = scores.get(key, {}).get('score', 0)
            cell = ws1[f'{col_letter}{row}']
            cell.value = s
            cell.fill = score_fill(s)
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # R, S, T = Interview/Reference — leave blank
        for col_letter in ['R', 'S', 'T']:
            ws1[f'{col_letter}{row}'].border = THIN_BORDER

        # U, V = Military/Foster preference
        for col_letter in ['U', 'V']:
            cell = ws1[f'{col_letter}{row}']
            if idx == 0:
                cell.value = 'Y or N'
            cell.fill = YELLOW_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # W = TOTAL SCORE (same as application score initially)
        cell = ws1[f'W{row}']
        cell.value = app_score
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Auto-fit column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 4
    for col_letter in [get_column_letter(i) for i in range(3, 24)]:
        ws1.column_dimensions[col_letter].width = 8

    # ── Sheet 2: Justifications ──
    ws2 = wb.create_sheet("Justifications")
    headers2 = ['Candidate Name', 'Criterion', 'Score', 'Justification',
                 'Resume Section', 'Keywords Found', 'Keywords Missing']
    for i, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = WHITE_FONT
        cell.fill = BLACK_FILL
        cell.border = THIN_BORDER

    row = 2
    for c in sorted(candidates, key=lambda x: (x['candidate_name'] or x['filename'])):
        scores = json.loads(c['scores_json']) if c['scores_json'] else {}
        name = c['candidate_name'] or c['filename']
        for key in CRITERIA_ORDER:
            s = scores.get(key, {})
            ws2.cell(row=row, column=1, value=name).border = THIN_BORDER
            ws2.cell(row=row, column=2, value=CRITERIA_LABELS.get(key, key)).border = THIN_BORDER
            score_cell = ws2.cell(row=row, column=3, value=s.get('score', 0))
            score_cell.border = THIN_BORDER
            score_cell.fill = score_fill(s.get('score', 0))
            just_cell = ws2.cell(row=row, column=4, value=s.get('justification', ''))
            just_cell.border = THIN_BORDER
            just_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws2.cell(row=row, column=5, value=s.get('resume_section', '')).border = THIN_BORDER
            ws2.cell(row=row, column=6, value=', '.join(s.get('keywords_found', []))).border = THIN_BORDER
            ws2.cell(row=row, column=7, value=', '.join(s.get('keywords_missing', []))).border = THIN_BORDER
            row += 1

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 100
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 40
    ws2.column_dimensions['G'].width = 40

    # ── Sheet 3: Executive Summary ──
    ws3 = wb.create_sheet("Executive Summary")
    headers3 = ['Rank', 'Candidate Name', 'Email', 'Application Score',
                 'Recommendation', 'Overall Summary', 'Top Strengths', 'Red Flags']
    for i, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=i, value=h)
        cell.font = WHITE_FONT
        cell.fill = BLACK_FILL
        cell.border = THIN_BORDER

    for idx, c in enumerate(candidates):
        row = idx + 2
        name = c['candidate_name'] or c['filename']
        strengths = json.loads(c['top_strengths']) if c['top_strengths'] else []
        flags = json.loads(c['red_flags']) if c['red_flags'] else []

        ws3.cell(row=row, column=1, value=idx + 1).border = THIN_BORDER
        ws3.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws3.cell(row=row, column=3, value=c['candidate_email'] or '').border = THIN_BORDER
        score_cell = ws3.cell(row=row, column=4, value=c['application_score'] or 0)
        score_cell.border = THIN_BORDER
        score_cell.font = Font(bold=True)
        ws3.cell(row=row, column=5, value=c.get('recommendation', '') or '').border = THIN_BORDER
        ws3.cell(row=row, column=6, value=c['overall_summary'] or '').border = THIN_BORDER
        ws3.cell(row=row, column=7, value='\n'.join(strengths)).border = THIN_BORDER
        ws3.cell(row=row, column=8, value='\n'.join(flags)).border = THIN_BORDER

    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 80
    ws3.column_dimensions['F'].width = 50
    ws3.column_dimensions['G'].width = 50

    # Save
    filename = 'ATS_Scoring_Matrix.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)
    return filepath, filename
